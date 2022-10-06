from json import load
from openpyxl import Workbook, load_workbook

def newline():
  return print()

# Creating a worksheet from the workbook
wb = Workbook()
# ws = wb.active # For getting the current worksheet
ws_01 = wb.create_sheet("01_Worksheet")

ws_01.title = "Test Sheet Excel 01"
# ws_02 = wb[${title}] # For getting the worksheet into a variable

# Setting tab color (under the worksheet)
ws_01.sheet_properties.tabColor = "000000"

# print(f"Worksheets: {wb.worksheets}")
for sheet in wb:
  print(sheet)

# Copying worksheet
source = wb.active
target = wb.copy_worksheet(source)

# Accessing cells
ws_01['A1'] = 'Names'
ws_01['B1'] = 'Rank'

C1 = ws_01.cell(row=1, column=3, value='Dept')

# Accessing the cells will create em in memory even if not assigned any value

# Accessing many cells
cell_range = ws_01['A1':'C2']

for row in ws_01.iter_rows(min_row=1, max_col=3, max_row=2, values_only=False):
  for cell in row:
    print(cell)

# Worksheet.rows or Worksheet.columns returns the values in tuples

# Getting the values in the worksheet
for row in ws_01.values:
  a = ''
  for value in row:
    a += str(value) + '\t'
  print(a)
newline()

# Retreiving cells data
print(C1.value)

# Saving the worksheet - it will overwrite existing file
wb.save("excel/01_worksheet.xlsx")

# Saving to a stream for webapp usage requires additional module
# from tempfile import NamedTemporaryFile
# wb = Workbook()
# with NamedTemporaryFile() as tmp:
#   wb.save(tmp.name)
#   tmp.seek(0)
#   stream = tmp.read()

# Saving worksheet as template
wb = load_workbook('excel/01_worksheet.xlsx')
wb.template = True
wb.save('excel/01_worksheet.xltx')

wb = load_workbook('excel/01_worksheet.xltx')
wb.template = False
wb.save('excel/01_worksheet.xlsx')

# Loading workbook
from openpyxl import load_workbook
wb2 = load_workbook('excel/02_worksheet.xlsx')
print(wb2.sheetnames)