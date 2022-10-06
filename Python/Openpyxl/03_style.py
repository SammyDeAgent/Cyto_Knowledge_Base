from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, GradientFill, colors, NamedStyle
from time import time

def main():

  # Loading in the worksheet
  start = int(time() * 1000)
  print('Opening worksheet...')
  wb = load_workbook("excel/03_worksheet.xlsx")
  ws = wb['Openpyxl Operations']
  end = int(time() * 1000)
  print(f'Worksheet loaded! Elapsed: {end - start}ms')
  newline(1)

  # Stylings
  font = Font(
    name='Consolas',
    size=10,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF000000'
  )
  fill = PatternFill(
    fill_type='solid',
    fgColor='CCCCAA00',
    bgColor='FF000000'
  )
  alignment = Alignment(
    horizontal='general',
    vertical='bottom',
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=False,
    indent=0
  )
  protection = Protection(
    locked=True,
    hidden=False
  )

  # Applying
  ws['A1'].font = font
  ws['C3'].font = Font(color='FF88FF00', italic=True)

  ws['A2'].fill = fill
  ws['A4'].fill = GradientFill(type='linear',degree=0, stop=(colors.BLUE, colors.WHITE))

  line1 = Side(border_style='medium', color='BB135678')
  ws['B2'].border = Border(
    left=line1,
    right=line1,
    top=line1,
    bottom=line1
    # diagonal=line1
    # diagonalUp=True
  )

  ws['A3'].alignment = Alignment(horizontal='center',vertical='center')

  # Creating own style
  myStyle = NamedStyle(name='General_Style')
  myStyle.font = Font('Comic Sans')
  myStyle.border = Border(
    left=line1,
    right=line1,
    top=line1,
    bottom=line1
  )
  myStyle.fill = PatternFill(
    fill_type='solid',
    bgColor='51A4156F',
    fgColor='00134941'
  )

  wb.add_named_style(myStyle)
  ws['C5'] = 'General Kenobi'
  ws['C5'].style = 'General_Style'

  # Saving the worksheet - it will overwrite existing file
  print('Saving worksheet...')
  start = int(time() * 1000)
  wb.save("excel/04_worksheet.xlsx")
  end = int(time() * 1000)
  print(f'Worksheet saved! Elapsed: {end - start}ms')

def newline(amount):
    i = 0
    while(i < amount):
      print()
      i+=1

if __name__ == "__main__":
  main()