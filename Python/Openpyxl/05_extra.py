# Library Imports
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Color
from PIL import Image
from time import time

def main():
  wb = Workbook()
  ws = wb.active

  # Setting an image to excel
  print("Loading image...")
  start = int(time() * 1000)
  img = Image.open('img/engie_nope.jpg')
  (image_width, image_height) = img.size
  pix = img.load()
  end = int(time() * 1000)
  print(f'Image Loaded! Elapsed: {end - start}ms')
  newline(1)

  start = int(time() * 1000)
  print('Generating Image...')
  for row in progressBar(range(1, image_height), prefix = 'Progress:', suffix = 'Complete', length = 50):
    for col in range(1, image_width):
      cell = ws.cell(column=col, row=row, value='')
      pixColor = "FF%02X%02X%02X" % (pix[col-1,row-1][0], pix[col-1,row-1][1],pix[col-1,row-1][2])
      cell.fill=PatternFill(patternType="solid",fgColor=Color(rgb=pixColor))
    ws.row_dimensions[row].height=3

  for col in range(1, image_width):
    ws.column_dimensions[get_column_letter(col)].width=0.5
  end = int(time() * 1000)
  print(f'Image generated! Elapsed: {(end - start)/1000}s')
  newline(1)
  
  # Saving the worksheet - it will overwrite existing file
  print('Saving worksheet...')
  start = int(time() * 1000)
  wb.save("excel/engie_nope.xlsx")
  end = int(time() * 1000)
  print(f'Worksheet saved! Elapsed: {(end - start)/1000}s')

  # Tips: get_column_letter(index) to get the row letters

def newline(amount):
    i = 0
    while(i < amount):
      print()
      i+=1

def progressBar(iterable, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r"):
    """
    Call in a loop to create terminal progress bar
    @params:
        iterable    - Required  : iterable object (Iterable)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """
    total = len(iterable)
    # Progress Bar Printing Function
    def printProgressBar (iteration):
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Initial Call
    printProgressBar(0)
    # Update Progress Bar
    for i, item in enumerate(iterable):
        yield item
        printProgressBar(i + 1)
    # Print New Line on Complete
    print()

if __name__ == "__main__":
  main()