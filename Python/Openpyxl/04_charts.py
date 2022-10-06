# Library Imports
from msilib.schema import tables
from sqlite3 import Date
from colorama import Style
from openpyxl import Workbook,load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart3D,
    PieChart,
    LineChart,
    ProjectedPieChart,
    Reference,
    Series,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart.series import DataPoint
from openpyxl.chart.axis import DateAxis
from openpyxl.worksheet.table import (
  Table,
  TableStyleInfo
)
from time import time

def main():

  # Loading in the worksheet
  start = int(time() * 1000)
  print('Opening worksheet...')
  wb = load_workbook("excel/05_worksheet.xlsx")
  ws = wb['Sheet1']
  end = int(time() * 1000)
  print(f'Worksheet loaded! Elapsed: {end - start}ms')
  newline(1)

  # Drawing a area chart
  chart = AreaChart()

  chart.title = 'Area Chart 1'
  chart.style = 13
  chart.x_axis.title = 'Test'
  chart.y_axis.title = 'Percentage'

  cats = Reference(ws, min_col=1, min_row=1, max_row=7)
  data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=7)

  chart.add_data(data, titles_from_data=True)
  chart.set_categories(cats)

  ws.add_chart(chart, 'A10')

  # Bar Chart
  ws = wb['Sheet2']

  chart = BarChart3D()
  chart.title = "3D Bar Chart"

  data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=4)
  cats = Reference(ws, min_col=1, min_row=2, max_row=4)

  chart.add_data(data=data, titles_from_data=True)
  chart.set_categories(cats)

  ws.add_chart(chart, "A7")

  # Piechart
  ws = wb['Sheet3']

  pie = PieChart()
  pie.title = "Weapons sold by category"

  cats = Reference(ws, min_col=1, min_row=2, max_row=5)
  data = Reference(ws, min_col=2, min_row=1, max_row=5)

  pie.add_data(data,titles_from_data=True)
  pie.set_categories(cats)

  # Cutting the first slice out
  slice = DataPoint(idx=0, explosion=20)
  pie.series[0].data_points = [slice]

  ws.add_chart(pie, "D1")

  # Linechart
  ws = wb['Sheet4']

  c2 = LineChart()
  c2.title = "Date Axis"
  c2.style = 12
  c2.x_axis = DateAxis(crossAx=100)
  c2.x_axis.number_format = 'd-mmm'
  c2.x_axis.title = 'Date'
  c2.y_axis.title = 'Size'
  c2.y_axis.crossAx = 500

  data = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=7)
  cats = Reference(ws, min_col=1, min_row=2, max_row=7)

  c2.add_data(data, titles_from_data=True)
  c2.set_categories(cats)

  ws.add_chart(c2, "A8")

  # Processing Table
  # ws = wb['Sheet5']
  # tab = Table(displayName='Table1', ref='A1:E5')
  # s1 = TableStyleInfo(
  #   name="TableStyleMedium9",
  #   showFirstColumn=False,
  #   showLastColumn=False,
  #   showRowStripes=True,
  #   showColumnStripes=False
  # )
  # tab.tableStyleInfo = s1
  # ws.add_table(tab)

  # Table Validation
  ws = wb['Sheet6']

  dv = DataValidation(type="list", formula1='"Dog, Cat, Bat"', allow_blank=True)
  ws.add_data_validation(dv)
  c1 = ws['A1']
  c1.value = 'Dog'
  dv.add(c1) # Can be in a range of cells

  # Saving the worksheet - it will overwrite existing file
  print('Saving worksheet...')
  start = int(time() * 1000)
  wb.save("excel/05_worksheet.xlsx")
  end = int(time() * 1000)
  print(f'Worksheet saved! Elapsed: {end - start}ms')

def newline(amount):
    i = 0
    while(i < amount):
      print()
      i+=1

if __name__ == "__main__":
  main()