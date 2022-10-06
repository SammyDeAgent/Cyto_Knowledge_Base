# Imports
from openpyxl import Workbook, load_workbook

def main():

  print('Opening worksheet...')
  wb = load_workbook("excel/02_worksheet.xlsx")
  ws = wb['Python Learning Material']
  print('Worksheet loaded!')
  newline(1)

  # Reading cells
  rows = ws.rows
  for row in rows:
    row_text = ''
    for cell in row:
      row_text += str(cell.value) + '\t'
    print(row_text)

  # For reading specific range of rows or columns would require proper indexing

  # Appending values
  ws2 = wb.create_sheet('Openpyxl Operations')
  list1 = ['Weapons','AK-47','M4A4','MP5','Kar98k','UMP-45']
  
  ws2.append(list1)
  
  # Dict can also be appended via index column label or just integer
  ws2.append(
    {
      'A':'Caliber',
      'B':'7.62x39',
      'C':'5.56x45',
      'D':'9x19',
      'E':'7.92x57',
      'F':'.45 acp',
    }
  )
  ws2.append(
    {
      1:  'Price',  
      2:  2700,
      3:  3300,
      4:  1500,
      5:  4750,
      6:  1200
    }
  )
  
  # Excel formulae operations - Refer to excel formulaes
  ws2['H3'] = 'Sum of Price'
  ws2['I3'] = '=SUM(B3:F3)'

  ws2['H4'] = 'Average of Price'
  ws2['I4'] = '=AVERAGE(B3:F3)'

  ws2['H1'] = 'Weapon Count'
  ws2['I1'] = '=COUNT(B3:F3)'

  ws2['H2'] = 'Is AK47 expensive than Kar98k?'
  ws2['I2'] = '=IF(B3>E3, "Yes", "No")'

  # Merging of cells - also can unmerge too
  ws2.merge_cells('A4:F4')

  # Grouping of rows - Ungroup level = 0
  for row in range(2,4):
    ws2.row_dimensions[row].outlineLevel = 1

  # Setting row and column attribute
  ws2.column_dimensions['A'].width = 20
  ws2.column_dimensions['H'].width = 50
  # ws2.row_dimensions['4'].hidden = True

  # Saving the worksheet - it will overwrite existing file
  print('Saving worksheet...')
  wb.save("excel/03_worksheet.xlsx")
  print('Worksheet saved!')

def newline(amount):
    i = 0
    while(i < amount):
      print()
      i+=1

if __name__ == "__main__":
  main()